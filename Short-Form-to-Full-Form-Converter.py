try:
	import keyboard
except:
	print('Please wait, we are installing keyboard library on your machine...')
	from pip._internal import main as pipmain
	pipmain(['install', 'keyboard'])
	try:
		import keyboard
	except:
		input("keyboard library not installed, run 'pip install keyboard' command in cmd, press enter to exit and run script again after installation")
		quit()
import pprint
choice=input('Enter\n1 to run program\n2 to load data from new excel file\n: ')
if choice=='1':
	import datafile
	input('\nprogram is running.......\nType short form then space to complete it to full form or press enter key to exit ')
if choice=='2':
	filename=input('data in excel file should be in given format, only .xlsx file type is supported. \nenter excel filename without extension: ')
	fp=open(filename+'.xlsx')
	try:
		import openpyxl
	except:
		print('Please wait, we are installing openpyxl library on your machine...')
		from pip._internal import main as pipmain
		pipmain(['install', 'openpyxl'])
		try:
			import openpyxl
		except:
			input("openpyxl library not installed, run 'pip install openpyxl' command in cmd, press enter to exit and run script again after installation")
			quit()
	excelfile=openpyxl.load_workbook(filename+'.xlsx')
	datasheet=excelfile.active
	print('loading',datasheet.max_row-1,'short forms')
	datadict={}
	for i in range(datasheet.max_row-1):
		datadict[str(datasheet.cell(row=2+i,column=2).value).strip()]=str(datasheet.cell(row=2+i,column=3).value).strip()
	mylist=list(datadict.keys())
	for short in mylist:
		print('%10s : %s' %(short,datadict[short]))
	fp1=open('datafile.py','w')
	fp1.write('import keyboard\ndatadict='+pprint.pformat(datadict)+'\n')
	fp1.write("mylist=list(datadict.keys())\n")
	fp1.write("for short in mylist:\n")
	fp1.write("\tprint('%10s : %s' %(short,datadict[short]))\n")
	for i in range(1,datasheet.max_row):
		fp1.write("keyboard.add_abbreviation('"+str(datasheet.cell(row=1+i,column=2).value).strip()+"', datadict['"+str(datasheet.cell(row=1+i,column=2).value).strip()+"'])\n")
	fp1.close()
	input("data loaded successfully, press enter to exit, run program again to use new short forms")
